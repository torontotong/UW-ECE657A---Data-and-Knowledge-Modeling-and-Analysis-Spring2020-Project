from itertools import cycle

import matplotlib.pyplot as plt
from numpy import *
import numpy as np
from os import listdir
import cv2
from sklearn.metrics import roc_curve, auc
from openpyxl import load_workbook
import csv


image_size = 512
train_dataset_folder = 'processed_data/'
test_dataset_folder = 'testset/'

def plot_roc(fpr, tpr, algorithm_num, type='pca'):
    # Compute ROC curve and ROC area for each class
    algm_name = ['SVM', 'Logistic Regression', 'AdaBoost']
    roc_auc = dict()
    thresholds = dict()
    for i in range(algorithm_num):
        roc_auc[i] = auc(fpr[i], tpr[i])

    lw = 1
    colors = cycle(['aqua', 'darkorange', 'cornflowerblue','red', 'green', 'blue'])
    for i, color in zip(range(algorithm_num), colors):
        plt.plot(fpr[i], tpr[i], color=color, lw=lw,
                 label='ROC of {0} (AUC = {1:0.2f})'
                       ''.format(algm_name[i], roc_auc[i]))
    plt.plot([0, 2], [0, 2], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver operating characteristic')
    plt.legend(loc="best")
    plt.savefig(type + "-roc-curve-" + str(image_size)+".png")
    plt.close()


def img2vector(filename):
    print('Load ' + filename)
    img = cv2.imread(filename, cv2.IMREAD_GRAYSCALE)
    res = cv2.resize(img, dsize=(1440,2880), interpolation=cv2.INTER_CUBIC)
    mat = np.array(res)
    k = 66
    ret_mat = mat.flatten()
    return ret_mat, k

def loadTestDataset():
    testingData = []
    testDataFileList = []
    testFileList = listdir(test_dataset_folder)
    m = len(testFileList)
    for i in range(m):
        fileNameStr = testFileList[i]
        fileStr = fileNameStr.split('.')[0]
        if fileStr =='':
            continue
        testDataMat, k= img2vector('%s/%s' % (test_dataset_folder, fileNameStr))
        testDataFileList.append(fileStr)
        testingData.append(testDataMat)
    return testingData, testDataFileList

def loadDataset(folder_name):
    print("Dataset Loading... ")
    class_labels = []
    training_data = []
    n_component = []
    classlist = listdir(folder_name)
    class_num = len(classlist)
    for class_idx in range(class_num):
        class_label = classlist[class_idx]
        if class_label == '.DS_Store':
            continue
        loadImages(class_label, training_data, class_labels, n_component)
    max_components = np.array(n_component).max()
    return training_data, class_labels , max_components

def loadImages(class_label, trainingData, labels, min_component):
    dirName = train_dataset_folder + class_label
    trainingFileList = listdir(dirName)
    m = len(trainingFileList)
    for i in range(m):
        fileNameStr = trainingFileList[i]
        fileStr = fileNameStr.split('.')[0]
        if fileStr =='':
            continue
        if class_label == 'uncomp':
            labels.append(0)
        elif class_label == 'comp':
            labels.append(1)
        trainingMat, k= img2vector('%s/%s' % (dirName, fileNameStr))
        trainingData.append(trainingMat)
        min_component.append(k)
    return

def get_confusion_matrix(y_test, predcted, class_num):
    my_confusion_matrix = np.zeros(shape=(class_num, class_num))
    for index in range(len(y_test)):
        if y_test[index] == predcted[index]:
            my_confusion_matrix[y_test[index]][y_test[index]] += 1
        else:
            my_confusion_matrix[y_test[index]][predcted[index]] += 1

    return my_confusion_matrix


def precision(label, cm):
    col = cm[:, label]
    return cm[label, label] / col.sum()

def recall(label, cm):
    row = cm[label, :]
    return cm[label, label] / row.sum()

def precision_average(cm):
    rows, columns = cm.shape
    sum_of_precisions = 0
    for row in range(rows):
        sum_of_precisions += precision(row, cm)
    return sum_of_precisions / rows

def recall_average(cm):
    rows, columns = cm.shape
    sum_of_recalls = 0
    for col in range(columns):
        sum_of_recalls += recall(col, cm)
    return sum_of_recalls / columns

def total_accuracy(cm):
    diagonal_sum = cm.trace()
    sum_of_all_elements = cm.sum()
    return diagonal_sum / sum_of_all_elements

def total_f1_score(cm):
    precision_factor = precision_average(cm)
    recall_factor = recall_average(cm)
    f1_factor = 2 * precision_factor * recall_factor / (precision_factor + recall_factor)
    return f1_factor

def write_excel(question,result):
    if question == 1:
        file_path = 'Dataset/Testing labels - Part I.xlsx'
    if question == 2:
        file_path = 'Dataset/Testing labels - Part II.xlsx'
    #Open an xlsx for reading
    wb = load_workbook(filename = file_path)
    ws = wb.get_sheet_by_name('Sheet1')
    for i in range(2, ws.max_row+1):
        img = ((ws.cell(row=i, column=1).value).split('.')[0]).split('_')[1]
        label_index = result[int(img)]
        if label_index == 0:
            ws.cell(row=i, column=2).value = 'Normal'
        elif label_index == 1:
            ws.cell(row=i, column=2).value = 'Pneumonia'
        elif label_index == 2:
            ws.cell(row=i, column=2).value = 'TB'
    #save the csb file
    wb.save(file_path)


#def main():
#    TrainingDataMat, Label_lists = loadDataset(train_dataset_folder)
#    return

